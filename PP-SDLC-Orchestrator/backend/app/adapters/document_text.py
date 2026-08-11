"""Extracts readable text from a generated artefact file, so a downstream
LLM adapter can pass an upstream agent's actual output as prompt context
instead of just re-sending the original task request at every phase.
"""

from pathlib import Path

from docx import Document
from openpyxl import load_workbook


def extract_text(file_path: Path, *, max_chars: int = 8000) -> str:
    """Best-effort plain-text extraction. Word docs are read paragraph by
    paragraph, Excel workbooks row by row per sheet (both binary zip
    formats — a raw text read would return garbled bytes, not content);
    anything else falls back to a raw read. Truncated so upstream context
    can't blow out a prompt budget.
    """
    suffix = file_path.suffix.lower()
    if suffix == ".docx":
        doc = Document(str(file_path))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    elif suffix == ".xlsx":
        wb = load_workbook(str(file_path), data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"[{sheet_name}]")
            for row in ws.iter_rows(values_only=True):
                values = [str(v) for v in row if v is not None and str(v).strip()]
                if values:
                    lines.append(" | ".join(values))
        text = "\n".join(lines)
    else:
        text = file_path.read_text(encoding="utf-8", errors="replace")

    return text[:max_chars]
