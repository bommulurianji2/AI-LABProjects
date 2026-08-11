import json

import pytest
from openpyxl import load_workbook

from app.adapters.llm_agent_adapter import TestAgentLlmAdapter
from app.adapters.model_providers.base import ModelProviderError
from app.agents_registry.contract import AgentRunRequest
from tests.fixtures.fake_model_provider import FakeModelProvider

_VALID_RESPONSE = {
    "test_cases": [
        {
            "type": "OQ",
            "description": "Verify the Request Form submits successfully with all required fields.",
            "related_entity": "REQ-001",
            "status": "Passed",
        },
        {
            "type": "SIT",
            "description": "Verify the reservation conflict-check flow rejects overlapping bookings.",
            "related_entity": "ADR-003",
            "status": "Failed",
        },
    ],
    "defects": [
        {
            "related_test": "TC-002",
            "description": "Conflict-check flow allows overlapping bookings under concurrent submission.",
            "status": "Open",
        },
    ],
}


def _make_request(**overrides) -> AgentRunRequest:
    defaults = dict(
        execution_mode="orchestrated",
        project_id="proj-1",
        invocation_id="inv-1",
        agent_id="test",
        task_id="task-1",
        task_request="Test the booking portal",
        lifecycle_phase="test",
        constraints={"project_name": "Facilities Booking Portal"},
        run_number=1,
        upstream_artefacts_text={
            "validation_report": "Pass with findings.",
            "solution_approach": "ADR-003: Deploy Power Automate cloud flows for booking confirmation.",
        },
    )
    defaults.update(overrides)
    return AgentRunRequest(**defaults)


def test_execute_renders_workbook_from_real_model_content():
    adapter = TestAgentLlmAdapter(provider=FakeModelProvider(json.dumps(_VALID_RESPONSE)))

    result = adapter.execute(_make_request())

    assert len(result.artefacts_produced) == 1
    produced = result.artefacts_produced[0]
    assert produced.artefact_type == "test_workbook"
    assert produced.entities == ["TC-001", "TC-002"]

    wb = load_workbook(produced.file_path)
    rows = list(wb["Test Cases"].iter_rows(values_only=True))
    assert rows[1] == ("TC-001", "OQ", _VALID_RESPONSE["test_cases"][0]["description"], "REQ-001", "Passed")
    assert rows[2][3] == "ADR-003"
    assert rows[2][4] == "Failed"

    defect_rows = list(wb["Defects"].iter_rows(values_only=True))
    assert defect_rows[1][0] == "DEF-001"
    assert defect_rows[1][1] == "TC-002"


def test_multiple_upstream_artefacts_are_included_in_prompt():
    provider = FakeModelProvider(json.dumps(_VALID_RESPONSE))
    adapter = TestAgentLlmAdapter(provider=provider)

    adapter.execute(_make_request())

    call = provider.calls[0]
    assert "Test Agent" in call["system"]
    assert "Pass with findings." in call["user"]
    assert "ADR-003: Deploy Power Automate" in call["user"]


def test_no_upstream_text_omits_context_gracefully():
    provider = FakeModelProvider(json.dumps(_VALID_RESPONSE))
    adapter = TestAgentLlmAdapter(provider=provider)

    adapter.execute(_make_request(upstream_artefacts_text={}))

    assert "Approved" not in provider.calls[0]["user"]


def test_case_without_related_entity_is_dropped():
    response = dict(_VALID_RESPONSE)
    response["test_cases"] = [
        {"type": "OQ", "description": "No traceability.", "related_entity": "", "status": "Passed"},
        {"type": "UAT", "description": "Has traceability.", "related_entity": "REQ-002", "status": "Passed"},
    ]
    adapter = TestAgentLlmAdapter(provider=FakeModelProvider(json.dumps(response)))

    result = adapter.execute(_make_request())

    assert result.artefacts_produced[0].entities == ["TC-001"]
    wb = load_workbook(result.artefacts_produced[0].file_path)
    rows = list(wb["Test Cases"].iter_rows(values_only=True))
    assert rows[1][2] == "Has traceability."


def test_missing_test_cases_raises_model_provider_error():
    response = dict(_VALID_RESPONSE)
    response["test_cases"] = []
    adapter = TestAgentLlmAdapter(provider=FakeModelProvider(json.dumps(response)))

    with pytest.raises(ModelProviderError):
        adapter.execute(_make_request())


def test_strips_model_supplied_related_entity_prefix():
    response = dict(_VALID_RESPONSE)
    response["test_cases"] = [
        {"type": "OQ", "description": "Test.", "related_entity": "REQ-1: REQ-001", "status": "Passed"}
    ]
    adapter = TestAgentLlmAdapter(provider=FakeModelProvider(json.dumps(response)))

    result = adapter.execute(_make_request())

    wb = load_workbook(result.artefacts_produced[0].file_path)
    rows = list(wb["Test Cases"].iter_rows(values_only=True))
    assert rows[1][3] == "REQ-001"


def test_malformed_json_raises_model_provider_error():
    adapter = TestAgentLlmAdapter(provider=FakeModelProvider("not json"))

    with pytest.raises(ModelProviderError):
        adapter.execute(_make_request())
