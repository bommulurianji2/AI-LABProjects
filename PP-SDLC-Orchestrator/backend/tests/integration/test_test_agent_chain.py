"""Extends the chain to an eighth real agent: ... -> Validation/QA -> Test.
Confirms the Test Workbook (Excel, not Word) is generated with correct
seeded TC-00N cases, sheet structure, and a working summary formula, and
that approval unlocks deploy.
"""

from openpyxl import load_workbook

from app.domain.enums import ArtefactVersionStatus, PhaseStatus
from app.models.artefact import ArtefactVersion
from app.models.user import User
from tests.helpers import make_orchestrator, run_phase_to_approval


def test_chain_through_test_agent_unlocks_deploy(db_session):
    orchestrator = make_orchestrator(db_session)
    reviewer = User(email="tester@example.test", role="Test Administrator")
    db_session.add(reviewer)
    db_session.commit()

    project = orchestrator.create_project("Contract and Approval Management")

    run_phase_to_approval(orchestrator, project, reviewer_id=reviewer.id, task_request="Draft requirements")
    run_phase_to_approval(orchestrator, project, reviewer_id=reviewer.id, task_request="Design the experience")
    run_phase_to_approval(
        orchestrator, project, reviewer_id=reviewer.id, task_request="Design the technical architecture"
    )
    run_phase_to_approval(orchestrator, project, reviewer_id=reviewer.id, task_request="Design the data model")
    run_phase_to_approval(
        orchestrator, project, reviewer_id=reviewer.id, task_request="Define governance and security controls"
    )
    run_phase_to_approval(orchestrator, project, reviewer_id=reviewer.id, task_request="Build the solution")
    run_phase_to_approval(
        orchestrator, project, reviewer_id=reviewer.id, task_request="Independently validate the build"
    )
    assert project.current_phase == "test"

    test_run = run_phase_to_approval(orchestrator, project, reviewer_id=reviewer.id, task_request="Execute test cases")
    assert project.current_phase == "deploy"
    assert project.phase_status == PhaseStatus.PENDING.value

    versions = db_session.query(ArtefactVersion).filter_by(run_id=test_run.id).all()
    assert len(versions) == 1
    version = versions[0]
    db_session.refresh(version)
    assert version.version_label == "v1.0"
    assert version.status == ArtefactVersionStatus.BASELINE.value
    assert version.file_path.endswith(".xlsx")

    wb = load_workbook(version.file_path)
    assert wb.sheetnames == ["Test Cases", "Summary", "Defects"]

    cases_ws = wb["Test Cases"]
    header = [c.value for c in cases_ws[1]]
    assert header == ["Test ID", "Type", "Description", "Related Entity", "Status"]

    data_rows = [tuple(c.value for c in row) for row in cases_ws.iter_rows(min_row=2)]
    assert len(data_rows) == 3
    assert data_rows[0][0] == "TC-001"
    assert all(row[4] == "Passed" for row in data_rows)

    summary_ws = wb["Summary"]
    total_formula = summary_ws["B2"].value
    assert total_formula == "=COUNTA('Test Cases'!A2:A1000)"

    defects_ws = wb["Defects"]
    assert [c.value for c in defects_ws[1]] == ["Defect ID", "Related Test", "Description", "Status"]
    assert defects_ws.max_row == 1  # header only, zero defects in this mock run
